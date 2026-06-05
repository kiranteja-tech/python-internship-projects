import tkinter as tk
from tkinter import filedialog, messagebox
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font
from openpyxl.drawing.image import Image
import os

# function : generate report
def generate_report(file_path):

    try:
        # read csv
        df = pd.read_csv(file_path)

        # remove hidden spaces in column names
        df.columns = df.columns.str.strip()

        # select numeric columns
        numeric_cols = df.select_dtypes(include='number')

        if numeric_cols.empty:
            messagebox.showerror("Error", "No numeric columns found!")
            return

        # summary statistics
        summary = pd.DataFrame({
            "Total": numeric_cols.sum(),
            "Average": numeric_cols.mean(),
            "Maximum": numeric_cols.max(),
            "Minimum": numeric_cols.min(),
            "Standard Deviation": numeric_cols.std()
        })

        # -------- Pivot Table (SAFE VERSION) --------
        df = df.loc[:, ~df.columns.duplicated()]

        cat_cols = df.select_dtypes(exclude='number').columns

        if len(cat_cols) == 0:
            messagebox.showerror("Error", "No category column found for pivot!")
            return

        index_col = cat_cols[0]
        value_col = numeric_cols.columns[0]

        pivot = pd.pivot_table(
            df,
            index=index_col,
            values=value_col,
            aggfunc="sum"
        )

        # Chart Creation 
        chart_path = "chart.png"

        pivot.plot(kind="bar")
        plt.title("Pivot Summary")
        plt.tight_layout()
        plt.savefig(chart_path)
        plt.close()

        # ASK USER WHERE TO SAVE 
        output_file = filedialog.asksaveasfilename(
            title="Choose location to save report",
            defaultextension=".xlsx",
            initialfile="report.xlsx",
            filetypes=[("Excel File", "*.xlsx")]
        )

        # if user cancels
        if not output_file:
            return

        #  Create Excel Workbook 
        wb = Workbook()

        # Sheet 1: Raw Data
        ws1 = wb.active
        ws1.title = "Raw Data"

        for r in dataframe_to_rows(df, index=False, header=True):
            ws1.append(r)

        for cell in ws1[1]:
            cell.font = Font(bold=True)

        # Sheet 2: Summary
        ws2 = wb.create_sheet("Summary")

        for r in dataframe_to_rows(summary, index=True, header=True):
            ws2.append(r)

        for cell in ws2[1]:
            cell.font = Font(bold=True)

        # Sheet 3: Pivot Table
        ws3 = wb.create_sheet("Pivot Table")

        for r in dataframe_to_rows(pivot, index=True, header=True):
            ws3.append(r)

        for cell in ws3[1]:
            cell.font = Font(bold=True)

        # Sheet 4: Bar Chart
        ws4 = wb.create_sheet("Bar Chart")

        img = Image(chart_path)
        ws4.add_image(img, "E2")

        # -------- Save Excel --------
        wb.save(output_file)

        # -------- Auto Open Excel (NEW) --------
        os.startfile(output_file)

        messagebox.showinfo(
            "Success",
            f"Excel Report Generated!\nSaved at:\n{output_file}"
        )

    except Exception as e:
        messagebox.showerror("Error", str(e))


# Tkinter GUI
def select_file():
    file_path = filedialog.askopenfilename(
        filetypes=[("CSV Files", "*.csv")]
    )
    if file_path:
        generate_report(file_path)


# GUI Window
root = tk.Tk()
root.title("Excel Report Generator")
root.geometry("400x200")

label = tk.Label(
    root,
    text="Select CSV File to Generate Report",
    font=("Arial", 12)
)
label.pack(pady=20)

btn = tk.Button(
    root,
    text="Choose CSV File",
    command=select_file,
    width=20,
    height=2
)
btn.pack()

root.mainloop()